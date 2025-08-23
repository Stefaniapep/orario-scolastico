#!/usr/bin/env python3
"""Generatore di orario conforme al README del progetto.

Il programma costruisce e ottimizza un modello CP-SAT per generare un orario scolastico,
minimizzando le ore di buco dei docenti e preferendo buchi di 2 ore.

Output: orario_settimanale.xlsx con fogli 'Classi' e 'Docenti', colorati per giorno.
"""

from ortools.sat.python import cp_model
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from collections import defaultdict
import math
import os

# --- 1. CONFIGURAZIONE DEI DATI DI INPUT ---
# In questa sezione vengono definiti tutti i dati grezzi che descrivono il problema.

CLASSI= ["1A", "1B", "2A", "2B", "3A", "3B", "4A", "4B", "5A", "5B"]
GIORNI= ["LUN", "MAR", "MER", "GIO", "VEN"]
SLOT_1= [("8:00-9:00",1.0),("9:00-10:00",1.0),("10:00-11:00",1.0),("11:00-12:00",1.0),("12:00-13:00",1.0),("13:00-13:30",0.5)]
SLOT_2= [("8:00-9:00",1.0),("9:00-10:00",1.0),("10:00-11:00",1.0),("11:00-12:00",1.0),("12:00-13:00",1.0),("13:00-14:00",1.0)]
SLOT_3= [("8:00-9:00",1.0),("9:00-10:00",1.0),("10:00-11:00",1.0),("11:00-12:00",1.0),("12:00-13:00",1.0)]
ASSEGNAZIONE_SLOT= { 
    "1A": { "LUN":"SLOT_1", "MAR":"SLOT_1", "MER":"SLOT_1", "GIO":"SLOT_1", "VEN":"SLOT_3" }, 
    "1B": { "LUN":"SLOT_1", "MAR":"SLOT_1", "MER":"SLOT_1", "GIO":"SLOT_1", "VEN":"SLOT_3" }, 
    "2A": { "LUN":"SLOT_1", "MAR":"SLOT_1", "MER":"SLOT_1", "GIO":"SLOT_1", "VEN":"SLOT_3" }, 
    "2B": { "LUN":"SLOT_1", "MAR":"SLOT_1", "MER":"SLOT_1", "GIO":"SLOT_1", "VEN":"SLOT_3" }, 
    "3A": { "LUN":"SLOT_1", "MAR":"SLOT_1", "MER":"SLOT_1", "GIO":"SLOT_1", "VEN":"SLOT_3" }, 
    "3B": { "LUN":"SLOT_1", "MAR":"SLOT_1", "MER":"SLOT_1", "GIO":"SLOT_1", "VEN":"SLOT_3" }, 
    "4A":{ "LUN":"SLOT_2", "MAR":"SLOT_2", "MER":"SLOT_2", "GIO":"SLOT_2", "VEN":"SLOT_3" }, 
    "4B":{ "LUN":"SLOT_2", "MAR":"SLOT_2", "MER":"SLOT_2", "GIO":"SLOT_2", "VEN":"SLOT_3" }, 
    "5A":{ "LUN":"SLOT_2", "MAR":"SLOT_2", "MER":"SLOT_2", "GIO":"SLOT_2", "VEN":"SLOT_3" }, 
    "5B":{ "LUN":"SLOT_2", "MAR":"SLOT_2", "MER":"SLOT_2", "GIO":"SLOT_2", "VEN":"SLOT_3" }, 
}
ORE_SETTIMANALI_CLASSI= { "1A": 27, "1B": 27, "2A": 27, "2B": 27, "3A": 27, "3B": 27, "4A": 29, "4B": 29, "5A": 29, "5B": 29, }
MAX_ORE_SETTIMANALI_DOCENTI=22
ASSEGNAZIONE_DOCENTI= { 
    "ANGELINI": {"1A": 11, "1B":11}, 
    "DOCENTE1": {"1A": 11, "1B":11}, 
    "DOCENTE3": {"5A": 11, "5B": 11}, 
    "SABATELLI": {"2A": 9, "2B":9, "copertura":4}, 
    "SCHIAVONE": {"2A": 11, "2B":11}, 
    "MARANGI": {"3A": 10, "3B":10, "copertura":2}, 
    "SIMEONE": {"3A": 11, "3B":11}, 
    "PEPE": {"4A": 8, "4B":8, "copertura":6}, 
    "PALMISANO": {"4A": 10, "4B":10, "copertura":2}, 
    "ZIZZI": {"5A": 11, "5B": 11}, 
    "CICCIMARRA": {"2A": 3, "2B":3, "copertura":6}, 
    "MOTORIA": {"5A": 2, "4A": 2, "4B":2, "5B":2 }, 
    "DOCENTE2": {"5A": 3,"5B": 3,"4A": 4, "4B":4}, 
    "LEO": {"1A":2, "1B":2, "2A":2, "2B":2, "3A":2, "3B":2, "4A":2, "4B":2, "5A":2, "5B":2}, 
    "SAVINO": {"1A":2, "1B":2, "2A":2, "2B":2, "3A":3, "3B":3, "4A":3, "4B":3}, 
    "DOCENTE4": {"1A": 1, "1B":1, "3A": 1, "3B":1}, }

# --- GRUPPI PER VINCOLI SPECIFICI (de-commentare per attivare) ---
GROUP_DAILY_TWO_CLASSES = {"ANGELINI","DOCENTE1","DOCENTE3","SABATELLI","SCHIAVONE","MARANGI","SIMEONE","PEPE","PALMISANO","ZIZZI"}
LIMIT_ONE_PER_DAY_PER_CLASS = {"MOTORIA","SAVINO"}
ONLY_DAYS = {
    "MOTORIA": {"MAR", "GIO", "VEN"}
}
START_AT = {
    "SCHIAVONE": {"LUN": 9, "MAR": 9, "GIO": 9}
}
END_AT = {
    "ZIZZI": {"MER": 10},
    "PEPE": {"LUN": 10}
}

if 'ONLY_DAYS' not in globals(): ONLY_DAYS = {}
if 'START_AT' not in globals(): START_AT = {}
if 'END_AT' not in globals(): END_AT = {}
if 'LIMIT_ONE_PER_DAY_PER_CLASS' not in globals(): LIMIT_ONE_PER_DAY_PER_CLASS = set()
if 'GROUP_DAILY_TWO_CLASSES' not in globals(): GROUP_DAILY_TWO_CLASSES = set()


# --- 2. PRE-ELABORAZIONE E DEFINIZIONE STRUTTURE DATI ---
# In questa sezione, i dati grezzi vengono trasformati in strutture più adatte
# alla modellazione matematica.

UNIT = 0.5  # [STRUTTURA] Unità di tempo base. 0.5 = 30 minuti. Tutte le ore vengono convertite in queste unità.
def hours_to_units(h): return int(round(h / UNIT))
def units_to_hours(u): return u * UNIT
def get_scheduling_label(time_str): return time_str.split('-')[0] # [LOGICA] Estrae l'ora di inizio (es. "8:00") come etichetta unica per la schedulazione.

SLOT_MAP = {"SLOT_1": SLOT_1, "SLOT_2": SLOT_2, "SLOT_3": SLOT_3}
# [STRUTTURA] `class_slots` è un dizionario nidificato che contiene, per ogni classe e giorno,
# la lista degli slot disponibili con tre informazioni: (etichetta_schedulazione, etichetta_completa, durata_in_unità).
class_slots = {cl: {day: [(get_scheduling_label(t), t, hours_to_units(d)) for t,d in SLOT_MAP[ASSEGNAZIONE_SLOT[cl][day]]] for day in GIORNI} for cl in CLASSI}

all_full_labels = list(set(t for s in [SLOT_1, SLOT_2, SLOT_3] for t, _ in s))
# [STRUTTURA] `GLOBAL_SCHEDULING_TIMES`: Lista ordinata di tutte le etichette di schedulazione uniche (es. "8:00", "9:00", ...).
# Rappresenta la "griglia" temporale di base per tutto l'orario.
GLOBAL_SCHEDULING_TIMES = sorted(list(set(get_scheduling_label(t) for t in all_full_labels)), key=lambda s_label: int(s_label.split(':')[0]))
EXCEL_LABELS = { (day, s_label): f"{day}{i+1}" for day in GIORNI for i, s_label in enumerate(GLOBAL_SCHEDULING_TIMES)}

teachers = list(ASSEGNAZIONE_DOCENTI.keys())
allowed_teachers_per_class = defaultdict(list)
for t,assign in ASSEGNAZIONE_DOCENTI.items():
    for cl in assign:
        if cl != 'copertura': allowed_teachers_per_class[cl].append(t)

total_copertura_units = sum(hours_to_units(assign.get('copertura', 0)) for assign in ASSEGNAZIONE_DOCENTI.values())
copertura_slots = defaultdict(list)
if total_copertura_units > 0:
    copertura_time_options = ['9:00-10:00', '10:00-11:00', '11:00-12:00', '12:00-13:00']
    units_per_day = math.ceil(total_copertura_units / len(GIORNI)); remaining = total_copertura_units; time_idx = 0
    for day in GIORNI:
        units_today = min(units_per_day, remaining)
        while units_today > 0:
            unit = 2 if units_today >= 2 else 1; time_label = copertura_time_options[time_idx % len(copertura_time_options)]
            copertura_slots[day].append((get_scheduling_label(time_label), time_label, unit)); units_today -= unit; remaining -= unit; time_idx += 1
        if remaining <= 0: break

def prevalidate_data():
    errors = []
    # 1) verificare che il totale delle ore assegnate ai docenti copra le ore richieste per ogni classe
    for cl in CLASSI:
        total_assigned = 0
        for t, assign in ASSEGNAZIONE_DOCENTI.items():
            total_assigned += assign.get(cl, 0)
        required = ORE_SETTIMANALI_CLASSI.get(cl, 0)
        if total_assigned < required:
            errors.append(f"Classe {cl}: ore assegnate totali {total_assigned}h < richieste {required}h")
    # 2) verificare che per ogni docente le ore assegnate (lezioni + coperture) non superino il massimo
    for t, assign in ASSEGNAZIONE_DOCENTI.items():
        lesson_hours = sum(v for k,v in assign.items() if k != 'copertura')
        cov = assign.get('copertura', 0)
        total = lesson_hours + cov
        if total > MAX_ORE_SETTIMANALI_DOCENTI:
            errors.append(f"Docente {t}: ore totali assegnate {total}h > max settimanale {MAX_ORE_SETTIMANALI_DOCENTI}h")
    # Report e abort se necessario
    if errors:
        print('\nPREVALIDAZIONE DATI FALLITA:')
        for e in errors:
            print(' -', e)
        print('\nCorreggi i dati in ASSEGNAZIONE_DOCENTI o aumenta MAX_ORE_SETTIMANALI_DOCENTI e rilancia.')
        exit(1)
    else:
        print('Prevalidazione dati OK: assegnazioni coprono le richieste di classe e rispettano i massimi docenti.', flush=True)
    print('Prevalidazione dati OK.', flush=True)

prevalidate_data()

# --- 3. CREAZIONE DEL MODELLO E DELLE VARIABILI ---
model = cp_model.CpModel()

# [STRUTTURA] `x` è la variabile di decisione principale.
# x[(classe, giorno, slot_index, docente)] è una BoolVar che è VERA se il docente `t`
# insegna nella classe `cl` durante lo slot `s_idx` del giorno `d`.
x = { (cl, day, s_idx, t): model.NewBoolVar(f"x_{cl}_{day}_{s_idx}_{t}") for cl in CLASSI for day in GIORNI for s_idx, _ in enumerate(class_slots[cl][day]) for t in allowed_teachers_per_class[cl] }

# [STRUTTURA] `copertura_vars` è la variabile di decisione per le ore di copertura.
copertura_vars = { (day, s_idx, t): (model.NewBoolVar(f"cop_{day}_{s_idx}_{t}"), sl, fl, u) for day, slots in copertura_slots.items() for s_idx, (sl, fl, u) in enumerate(slots) for t in teachers if ASSEGNAZIONE_DOCENTI.get(t, {}).get('copertura', 0) > 0 }

# [STRUTTURA] `b` (busy) è una variabile ausiliaria.
# b[(docente, giorno, ora_schedulazione)] è VERA se il docente `t` è impegnato
# (in classe o in copertura) durante lo slot che inizia all'ora `sched_label` del giorno `d`.
# Serve a semplificare i vincoli di sovrapposizione e continuità.
b = { (t, day, sched_label): model.NewBoolVar(f"b_{t}_{day}_{sched_label}") for t in teachers for day in GIORNI for sched_label in GLOBAL_SCHEDULING_TIMES }


# --- 4. APPLICAZIONE DEI VINCOLI FONDAMENTALI ---
# Questi vincoli sono le regole "dure" che ogni soluzione valida deve rispettare.

# [VINCOLO] Collega le variabili di assegnazione (x, copertura_vars) alla variabile di presenza (b).
# Garantisce anche che un docente non possa essere in due posti contemporaneamente.
for t, day, sched_label in b:
    vars_at_time = [x.get((cl, day, s_idx, t)) for cl in CLASSI for s_idx, (s_label, fl, u) in enumerate(class_slots[cl][day]) if s_label == sched_label]
    vars_at_time.extend(var for (d, s_idx, tt), (var, s_label, fl, u) in copertura_vars.items() if d == day and tt == t and s_label == sched_label)
    vars_at_time = [v for v in vars_at_time if v is not None]
    model.Add(sum(vars_at_time) <= 1)
    if vars_at_time:
        model.Add(sum(vars_at_time) >= 1).OnlyEnforceIf(b[(t, day, sched_label)])
        model.Add(sum(vars_at_time) == 0).OnlyEnforceIf(b[(t, day, sched_label)].Not())
    else: model.Add(b[(t, day, sched_label)] == 0)

# [VINCOLO] Ogni slot orario di una classe deve avere esattamente un docente.
for cl in CLASSI:
    for day in GIORNI:
        for s_idx, _ in enumerate(class_slots[cl][day]): model.Add(sum(x[(cl, day, s_idx, t)] for t in allowed_teachers_per_class[cl]) == 1)

# [VINCOLO] Il totale delle ore settimanali per ogni classe deve corrispondere a quanto definito.
for cl in CLASSI:
    model.Add(sum(x[(cl, day, s_idx, t)] * u for day in GIORNI for s_idx, (sl, fl, u) in enumerate(class_slots[cl][day]) for t in allowed_teachers_per_class[cl]) == hours_to_units(ORE_SETTIMANALI_CLASSI[cl]))

# [VINCOLO] Il totale delle ore di un docente in una specifica classe deve corrispondere a quanto definito.
for t, assign in ASSEGNAZIONE_DOCENTI.items():
    for cl, hours in assign.items():
        if cl != 'copertura': model.Add(sum(x.get((cl, day, s_idx, t), 0) * u for day in GIORNI for s_idx, (sl, fl, u) in enumerate(class_slots[cl][day])) == hours_to_units(hours))

# [VINCOLO] Gestione delle ore di copertura.
if total_copertura_units > 0:
    # Ogni slot di copertura deve essere assegnato a un solo docente.
    for day, slots in copertura_slots.items():
        for s_idx, _ in enumerate(slots): model.Add(sum(copertura_vars[(day, s_idx, t)][0] for t in teachers if (day, s_idx, t) in copertura_vars) == 1)
    # Ogni docente deve fare il suo totale di ore di copertura.
    for t in teachers:
        needed = hours_to_units(ASSEGNAZIONE_DOCENTI.get(t, {}).get('copertura', 0))
        model.Add(sum(var * u for (d, s_idx, tt), (var, sl, fl, u) in copertura_vars.items() if tt == t) == needed)


# --- 5. VINCOLI DI QUALITA' E SPECIFICI ---
# Questi vincoli definiscono le "preferenze" e le regole speciali.
print("\nApplicazione vincoli...")
active_constraints_for_report = []

print("- Vincolo: Massimo 4 ore per docente per classe al giorno")
for t, assignments in ASSEGNAZIONE_DOCENTI.items():
    for cl in assignments:
        if cl == 'copertura': continue
        for day in GIORNI:
            daily_class_units = sum(x.get((cl, day, s_idx, t), 0) * u for s_idx, (sl, fl, u) in enumerate(class_slots[cl][day]))
            model.Add(daily_class_units <= hours_to_units(4))

if LIMIT_ONE_PER_DAY_PER_CLASS:
    active_constraints_for_report.append(f"Max 1 ora/giorno/classe per {LIMIT_ONE_PER_DAY_PER_CLASS}")
    for t in LIMIT_ONE_PER_DAY_PER_CLASS:
        for cl in CLASSI:
            for day in GIORNI: model.Add(sum(x.get((cl, day, s_idx, t), 0) * u for s_idx, (sl, fl, u) in enumerate(class_slots[cl][day])) <= hours_to_units(1))
if ONLY_DAYS:
    active_constraints_for_report.append(f"Regole di giorni consentiti per {list(ONLY_DAYS.keys())}")
    for teacher, allowed_days in ONLY_DAYS.items():
        for day in set(GIORNI) - allowed_days:
            for sched_label in GLOBAL_SCHEDULING_TIMES:
                if (teacher, day, sched_label) in b: model.Add(b[(teacher, day, sched_label)] == 0)
if GROUP_DAILY_TWO_CLASSES:
    active_constraints_for_report.append(f"Almeno 1h/giorno in entrambe le classi per {GROUP_DAILY_TWO_CLASSES}")
    for t in GROUP_DAILY_TWO_CLASSES:
        classes = [c for c in ASSEGNAZIONE_DOCENTI.get(t, {}) if c != 'copertura']
        if len(classes) == 2:
            for day in GIORNI:
                for cl in classes: model.Add(sum(x.get((cl, day, s_idx, t), 0) * u for s_idx, (sl, fl, u) in enumerate(class_slots[cl][day])) >= hours_to_units(1))
if START_AT:
    active_constraints_for_report.append(f"Regole di inizio orario per {list(START_AT.keys())}")
    for teacher, rules in START_AT.items():
        for day, start_hour in rules.items():
            for sched_label in GLOBAL_SCHEDULING_TIMES:
                if int(sched_label.split(':')[0]) < start_hour:
                    if (teacher, day, sched_label) in b: model.Add(b[(teacher, day, sched_label)] == 0)
if END_AT:
    active_constraints_for_report.append(f"Regole di fine orario per {list(END_AT.keys())}")
    for teacher, rules in END_AT.items():
        for day, end_hour in rules.items():
            for sched_label in GLOBAL_SCHEDULING_TIMES:
                if int(sched_label.split(':')[0]) >= end_hour:
                    if (teacher, day, sched_label) in b: model.Add(b[(teacher, day, sched_label)] == 0)

print("- Vincolo: Blocchi di 2 o 3 ore in una classe devono essere consecutivi")
for t in teachers:
    if t in LIMIT_ONE_PER_DAY_PER_CLASS: continue
    for cl in ASSEGNAZIONE_DOCENTI.get(t, {}):
        if cl == 'copertura': continue
        for day in GIORNI:
            daily_class_units = sum(x.get((cl, day, s_idx, t), 0) * u for s_idx, (sl, fl, u) in enumerate(class_slots[cl][day]))
            
            # [MODIFICA] Esteso il vincolo per 2 E 3 ore
            is_2_or_3_hours = model.NewBoolVar(f'is_2_or_3h_{t}_{cl}_{day}')
            is_exactly_two = model.NewBoolVar(f'exactly_2h_{t}_{cl}_{day}')
            is_exactly_three = model.NewBoolVar(f'exactly_3h_{t}_{cl}_{day}')
            model.Add(daily_class_units == hours_to_units(2)).OnlyEnforceIf(is_exactly_two)
            model.Add(daily_class_units != hours_to_units(2)).OnlyEnforceIf(is_exactly_two.Not())
            model.Add(daily_class_units == hours_to_units(3)).OnlyEnforceIf(is_exactly_three)
            model.Add(daily_class_units != hours_to_units(3)).OnlyEnforceIf(is_exactly_three.Not())
            model.AddBoolOr([is_exactly_two, is_exactly_three]).OnlyEnforceIf(is_2_or_3_hours)
            model.AddImplication(is_2_or_3_hours.Not(), is_exactly_two.Not())
            model.AddImplication(is_2_or_3_hours.Not(), is_exactly_three.Not())

            teaches_this_class = []
            for sl in GLOBAL_SCHEDULING_TIMES:
                presence_in_slot = model.NewBoolVar(f'presence_{t}_{cl}_{day}_{sl.replace(":", "")}')
                vars_in_slot = [x.get((cl, day, s_idx, t)) for s_idx, (s_label, _, _) in enumerate(class_slots[cl][day]) if s_label == sl]
                if not vars_in_slot: model.Add(presence_in_slot == 0)
                else:
                    model.Add(sum(vars_in_slot) >= 1).OnlyEnforceIf(presence_in_slot)
                    model.Add(sum(vars_in_slot) == 0).OnlyEnforceIf(presence_in_slot.Not())
                teaches_this_class.append(presence_in_slot)
            
            starts = [model.NewBoolVar(f'class_start_{t}_{cl}_{day}_{i}') for i in range(len(teaches_this_class))]
            model.Add(starts[0] == teaches_this_class[0])
            for i in range(1, len(teaches_this_class)):
                model.AddBoolAnd([teaches_this_class[i], teaches_this_class[i-1].Not()]).OnlyEnforceIf(starts[i])
                model.AddBoolOr([starts[i], teaches_this_class[i].Not(), teaches_this_class[i-1]])
            
            num_class_blocks = sum(starts)
            model.Add(num_class_blocks <= 1).OnlyEnforceIf(is_2_or_3_hours)

print("- Vincolo: Continuità oraria flessibile (max 1 buco) per tutti i docenti")
for t in teachers:
    for day in GIORNI:
        works_at_time = [b[(t, day, sched_label)] for sched_label in GLOBAL_SCHEDULING_TIMES]
        starts = [model.NewBoolVar(f'start_{t}_{day}_{i}') for i in range(len(GLOBAL_SCHEDULING_TIMES))]
        model.Add(starts[0] == works_at_time[0])
        for i in range(1, len(GLOBAL_SCHEDULING_TIMES)):
            model.AddBoolAnd([works_at_time[i], works_at_time[i-1].Not()]).OnlyEnforceIf(starts[i])
            model.AddBoolOr([starts[i], works_at_time[i].Not(), works_at_time[i-1]])
        model.Add(sum(starts) <= 2)

# --- 6. OBIETTIVO DI OTTIMIZZAZIONE (MINIMIZZAZIONE BUCHI) ---
# [OTTIMIZZAZIONE] In questa sezione si definisce cosa rende un orario "migliore" di un altro.
# L'obiettivo è minimizzare una penalità calcolata sui buchi orari.

holes = {} # [STRUTTURA] h[(docente, giorno, ora)] è VERA se il docente ha un buco in quello slot.
for t in teachers:
    for day in GIORNI:
        works_at_time = [b[(t, day, sl)] for sl in GLOBAL_SCHEDULING_TIMES]
        
        has_worked_before = [model.NewBoolVar(f'hwb_{t}_{day}_{i}') for i in range(len(GLOBAL_SCHEDULING_TIMES))]
        model.Add(has_worked_before[0] == 0)
        for i in range(1, len(GLOBAL_SCHEDULING_TIMES)):
            model.AddBoolOr([has_worked_before[i-1], works_at_time[i-1]]).OnlyEnforceIf(has_worked_before[i])
            model.AddImplication(has_worked_before[i].Not(), has_worked_before[i-1].Not())
            model.AddImplication(has_worked_before[i].Not(), works_at_time[i-1].Not())

        will_work_after = [model.NewBoolVar(f'wwa_{t}_{day}_{i}') for i in range(len(GLOBAL_SCHEDULING_TIMES))]
        model.Add(will_work_after[-1] == 0)
        for i in range(len(GLOBAL_SCHEDULING_TIMES) - 2, -1, -1):
            model.AddBoolOr([will_work_after[i+1], works_at_time[i+1]]).OnlyEnforceIf(will_work_after[i])
            model.AddImplication(will_work_after[i].Not(), will_work_after[i+1].Not())
            model.AddImplication(will_work_after[i].Not(), works_at_time[i+1].Not())

        for i, sl in enumerate(GLOBAL_SCHEDULING_TIMES):
            h = model.NewBoolVar(f'h_{t}_{day}_{i}')
            model.AddBoolAnd([works_at_time[i].Not(), has_worked_before[i], will_work_after[i]]).OnlyEnforceIf(h)
            model.AddBoolOr([h, works_at_time[i], has_worked_before[i].Not(), will_work_after[i].Not()])
            holes[(t, day, sl)] = h
            
total_penalty = []
for t in teachers:
    for day in GIORNI:
        daily_hole_units = sum(holes[(t, day, sl)] for sl in GLOBAL_SCHEDULING_TIMES)
        is_zero_holes = model.NewBoolVar(f'is_zero_h_{t}_{day}'); model.Add(daily_hole_units == 0).OnlyEnforceIf(is_zero_holes); model.Add(daily_hole_units != 0).OnlyEnforceIf(is_zero_holes.Not())
        is_two_hour_hole = model.NewBoolVar(f'is_2h_h_{t}_{day}'); model.Add(daily_hole_units == hours_to_units(2)).OnlyEnforceIf(is_two_hour_hole); model.Add(daily_hole_units != hours_to_units(2)).OnlyEnforceIf(is_two_hour_hole.Not())
        is_good_hole_day = model.NewBoolVar(f'is_good_h_day_{t}_{day}'); model.AddBoolOr([is_zero_holes, is_two_hour_hole]).OnlyEnforceIf(is_good_hole_day)
        model.AddImplication(is_good_hole_day.Not(), is_zero_holes.Not()); model.AddImplication(is_good_hole_day.Not(), is_two_hour_hole.Not())
        daily_penalty = model.NewIntVar(0, 1000, f'penalty_{t}_{day}')
        model.Add(daily_penalty == 0).OnlyEnforceIf(is_zero_holes)
        model.Add(daily_penalty == 1).OnlyEnforceIf(is_two_hour_hole)
        model.Add(daily_penalty == daily_hole_units * 10).OnlyEnforceIf(is_good_hole_day.Not())
        total_penalty.append(daily_penalty)

model.Minimize(sum(total_penalty))

# --- 7. RISOLUZIONE E OUTPUT ---
print(f"Vincoli specifici attivi: {active_constraints_for_report if active_constraints_for_report else ['Nessuno']}")
print("\nAvvio ottimizzazione modello (minimizzazione buchi)...")
solver = cp_model.CpSolver(); solver.parameters.max_time_in_seconds = 120; solver.parameters.num_search_workers=os.cpu_count() or 8;
solver.parameters.randomize_search = True
res = solver.Solve(model)

def run_diagnostics(solver, has_solution):
    if not has_solution:
        print("\n--- ANALISI DI FATTIBILITA' DEI VINCOLI ---")
        if active_constraints_for_report:
            print("Il modello è insolubile con i seguenti vincoli attivi:")
            for c in active_constraints_for_report: print(f"  - {c}")
            print("\nSUGGERIMENTO: Prova a disattivare i vincoli più restrittivi (es. START_AT, END_AT, GROUP_DAILY) uno alla volta per trovare il punto di conflitto.")
        else:
            print("Il modello è insolubile anche senza vincoli specifici. Controllare i dati di base (ore, assegnazioni).")
        return

    print("\n--- VERIFICA DEI VINCOLI SULLA SOLUZIONE TROVATA ---"); report = []
    
    # Verifica Vincoli Fondamentali
    is_ok_class_hours = True; details_class_hours = []
    for cl in CLASSI:
        required = hours_to_units(ORE_SETTIMANALI_CLASSI[cl])
        found = sum(solver.Value(x[(cl, day, s_idx, t)]) * u for day in GIORNI for s_idx, (_, _, u) in enumerate(class_slots[cl][day]) for t in allowed_teachers_per_class[cl])
        if required != found: is_ok_class_hours = False; details_class_hours.append(f"  - FAIL: Classe {cl} - Richieste: {units_to_hours(required)}h, Trovate: {units_to_hours(found)}h")
    report.append(f"[{'PASS' if is_ok_class_hours else 'FAIL'}] Ore settimanali totali per classe"); report.extend(details_class_hours)
    is_ok_teacher_assign = True; details_teacher_assign = []
    for t, assignments in ASSEGNAZIONE_DOCENTI.items():
        for cl, hours in assignments.items():
            if cl == 'copertura': continue
            required = hours_to_units(hours)
            found = sum(solver.Value(x.get((cl, day, s_idx, t), 0)) * u for day in GIORNI for s_idx, (_, _, u) in enumerate(class_slots[cl][day]))
            if required != found: is_ok_teacher_assign = False; details_teacher_assign.append(f"  - FAIL: Docente {t} in Classe {cl} - Richieste: {units_to_hours(required)}h, Trovate: {units_to_hours(found)}h")
    report.append(f"[{'PASS' if is_ok_teacher_assign else 'FAIL'}] Ore specifiche Docente-Classe"); report.extend(details_teacher_assign)
    
    # Verifica Vincoli di Qualità (sempre attivi)
    is_ok_max_daily = True; details_max_daily = []
    for t, assignments in ASSEGNAZIONE_DOCENTI.items():
        for cl in assignments:
            if cl == 'copertura': continue
            for day in GIORNI:
                found_units = sum(solver.Value(x.get((cl, day, s_idx, t), 0)) * u for s_idx, (_, _, u) in enumerate(class_slots[cl][day]))
                if found_units > hours_to_units(4): is_ok_max_daily = False; details_max_daily.append(f"  - FAIL: {t} in {cl} il {day} ha {units_to_hours(found_units)}h (> 4h).")
    report.append(f"[{'PASS' if is_ok_max_daily else 'FAIL'}] Massimo 4 ore/giorno per docente nella stessa classe"); report.extend(details_max_daily)
    max_blocks_found = 0; violations = []
    for t in teachers:
        for day in GIORNI:
            works_at_time = [solver.Value(b.get((t, day, sl), 0)) for sl in GLOBAL_SCHEDULING_TIMES]
            if not any(works_at_time): continue
            num_blocks = sum([works_at_time[0]] + [1 for i in range(1, len(works_at_time)) if works_at_time[i] and not works_at_time[i-1]])
            max_blocks_found = max(max_blocks_found, num_blocks)
            if num_blocks > 2: violations.append(f"  - FAIL: {t} il {day} ha {num_blocks-1} buchi.")
    report.append(f"[{'PASS' if max_blocks_found <= 2 else 'FAIL'}] Continuità oraria (max 1 buco): Max blocchi trovati: {max_blocks_found}."); report.extend(violations)

    # [MODIFICA] Estesa la verifica per 2 E 3 ore
    is_ok_consecutive = True; details_consecutive = []
    for t in teachers:
        if t in LIMIT_ONE_PER_DAY_PER_CLASS: continue
        for cl in ASSEGNAZIONE_DOCENTI.get(t, {}):
            if cl == 'copertura': continue
            for day in GIORNI:
                daily_units = sum(solver.Value(x.get((cl, day, s_idx, t), 0)) * u for s_idx, (_, _, u) in enumerate(class_slots[cl][day]))
                if daily_units in [hours_to_units(2), hours_to_units(3)]:
                    indices = [i for i, sl in enumerate(GLOBAL_SCHEDULING_TIMES) if any(solver.Value(x.get((cl, day, s_idx, t), 0)) for s_idx, (s_label, _, _) in enumerate(class_slots[cl][day]) if s_label == sl)]
                    if len(indices) > 0 and (max(indices) - min(indices) > len(indices) - 1):
                        is_ok_consecutive = False
                        taught_at = [GLOBAL_SCHEDULING_TIMES[i] for i in indices]
                        details_consecutive.append(f"  - FAIL: {t} in {cl} il {day} ha {units_to_hours(daily_units)} ore non consecutive ({', '.join(taught_at)}).")
    report.append(f"[{'PASS' if is_ok_consecutive else 'FAIL'}] Lezioni di 2 o 3 ore sono consecutive"); report.extend(details_consecutive)

    # Verifica Vincoli Specifici (se attivi)
    if LIMIT_ONE_PER_DAY_PER_CLASS:
        is_ok = True; details = []
        for t in LIMIT_ONE_PER_DAY_PER_CLASS:
            for cl, hours in ASSEGNAZIONE_DOCENTI.get(t, {}).items():
                if cl == 'copertura': continue
                for day in GIORNI:
                    units_taught = sum(solver.Value(x.get((cl, day, s_idx, t), 0)) * u for s_idx, (_, _, u) in enumerate(class_slots[cl][day]))
                    if units_taught > hours_to_units(1): is_ok = False; details.append(f"  - FAIL: {t} in {cl} il {day} ha {units_to_hours(units_taught)}h (> 1h)")
        report.append(f"[{'PASS' if is_ok else 'FAIL'}] Max 1 ora/giorno/classe per {LIMIT_ONE_PER_DAY_PER_CLASS}"); report.extend(details)
    if ONLY_DAYS:
        is_ok = True; details = []
        for teacher, allowed_days in ONLY_DAYS.items():
            for day in set(GIORNI) - allowed_days:
                work_on_forbidden_day = any(solver.Value(b.get((teacher, day, sl), 0)) for sl in GLOBAL_SCHEDULING_TIMES)
                if work_on_forbidden_day: is_ok = False; details.append(f"  - FAIL: {teacher} lavora il {day}, che non è un giorno consentito.")
        report.append(f"[{'PASS' if is_ok else 'FAIL'}] Regole di giorni consentiti per {list(ONLY_DAYS.keys())}"); report.extend(details)
    if GROUP_DAILY_TWO_CLASSES:
        is_ok = True; details = []
        for t in GROUP_DAILY_TWO_CLASSES:
            classes = [c for c in ASSEGNAZIONE_DOCENTI.get(t, {}) if c != 'copertura']
            if len(classes) == 2:
                for day in GIORNI:
                    for cl in classes:
                        units_taught = sum(solver.Value(x.get((cl, day, s_idx, t), 0)) * u for s_idx, (_, _, u) in enumerate(class_slots[cl][day]))
                        if units_taught < hours_to_units(1): is_ok = False; details.append(f"  - FAIL: {t} in {cl} il {day} ha solo {units_to_hours(units_taught)}h (richiesta >= 1h).")
        report.append(f"[{'PASS' if is_ok else 'FAIL'}] Almeno 1h/giorno in entrambe le classi per {GROUP_DAILY_TWO_CLASSES}"); report.extend(details)
    if START_AT:
        is_ok = True; details = []
        for teacher, rules in START_AT.items():
            for day, start_hour in rules.items():
                for sched_label in GLOBAL_SCHEDULING_TIMES:
                    if int(sched_label.split(':')[0]) < start_hour:
                        if solver.Value(b.get((teacher, day, sched_label), 0)): is_ok = False; details.append(f"  - FAIL: {teacher} lavora alle {sched_label} di {day}, violando la regola di inizio ore {start_hour}.")
        report.append(f"[{'PASS' if is_ok else 'FAIL'}] Regole di inizio orario per {list(START_AT.keys())}"); report.extend(details)
    if END_AT:
        is_ok = True; details = []
        for teacher, rules in END_AT.items():
            for day, end_hour in rules.items():
                for sched_label in GLOBAL_SCHEDULING_TIMES:
                    if int(sched_label.split(':')[0]) >= end_hour:
                        if solver.Value(b.get((teacher, day, sched_label), 0)): is_ok = False; details.append(f"  - FAIL: {teacher} lavora alle {sched_label} di {day}, violando la regola di fine ore {end_hour}.")
        report.append(f"[{'PASS' if is_ok else 'FAIL'}] Regole di fine orario per {list(END_AT.keys())}"); report.extend(details)

    # Report finale sull'ottimizzazione
    total_hole_hours = sum(units_to_hours(solver.Value(h)) for h in holes.values())
    non_2h_hole_days = 0
    for t in teachers:
        for day in GIORNI:
            daily_hole_units = sum(solver.Value(holes.get((t, day, sl), 0)) for sl in GLOBAL_SCHEDULING_TIMES)
            if daily_hole_units > 0 and daily_hole_units != hours_to_units(2):
                non_2h_hole_days += 1
    report.append(f"[INFO] Ottimizzazione buchi: Trovate {total_hole_hours} ore di buco totali.")
    if non_2h_hole_days > 0:
        report.append(f"  - ATTENZIONE: Ci sono {non_2h_hole_days} orari giornalieri con buchi di durata diversa da 2 ore.")
    else:
        report.append("  - OTTIMO: Tutti i buchi presenti sono di 0 o 2 ore.")
    
    for line in report: print(line)

if res in (cp_model.OPTIMAL, cp_model.FEASIBLE):
    print("\nSoluzione trovata. Generazione file Excel in corso...")
    wb = Workbook(); day_colors = {"LUN": "FFFFCC", "MAR": "CCFFCC", "MER": "CCE5FF", "GIO": "FFDDCC", "VEN": "E5CCFF"}
    
    ws_classi = wb.active; ws_classi.title = "Classi"; ws_classi.append(["Slot"] + CLASSI + ["Copertura"])
    orario_classi = defaultdict(dict); orario_copertura = defaultdict(str)
    for (cl, day, s_idx, t), var in x.items():
        if solver.Value(var) == 1: orario_classi[(day, class_slots[cl][day][s_idx][0])][cl] = t
    for (d, s_idx, t), (var, sl, fl, u) in copertura_vars.items():
        if solver.Value(var) == 1: orario_copertura[(d, sl)] += f"{t} "
    for day in GIORNI:
        for sched_label in GLOBAL_SCHEDULING_TIMES:
            row_data = [EXCEL_LABELS[(day, sched_label)]] + [orario_classi.get((day, sched_label), {}).get(cl, "") for cl in CLASSI] + [orario_copertura.get((day, sched_label), "").strip()]
            ws_classi.append(row_data)
            for cell in ws_classi[ws_classi.max_row]: cell.fill = PatternFill(start_color=day_colors[day], end_color=day_colors[day], fill_type="solid")

    ws_docenti = wb.create_sheet("Docenti"); ws_docenti.append(["Slot"] + teachers)
    orario_docenti = defaultdict(dict)
    for (cl, day, s_idx, t), var in x.items():
        if solver.Value(var) == 1: orario_docenti[(day, class_slots[cl][day][s_idx][0])][t] = cl
    for (d, s_idx, t), (var, sl, fl, u) in copertura_vars.items():
        if solver.Value(var) == 1: orario_docenti[(d, sl)][t] = "COPERTURA"
    for t in teachers:
        for day in GIORNI:
            if any(solver.Value(h) for sl, h in holes.items() if h.Name().startswith(f'h_{t}_{day}')):
                for sl in GLOBAL_SCHEDULING_TIMES:
                    if solver.Value(holes[(t, day, sl)]):
                        orario_docenti[(day, sl)][t] = "BUCO"

    for day in GIORNI:
        for sched_label in GLOBAL_SCHEDULING_TIMES:
            row_data = [EXCEL_LABELS[(day, sched_label)]] + [orario_docenti.get((day, sched_label), {}).get(t, "") for t in teachers]
            ws_docenti.append(row_data)
            for cell in ws_docenti[ws_docenti.max_row]: cell.fill = PatternFill(start_color=day_colors[day], end_color=day_colors[day], fill_type="solid")
    
    wb.save("orario_settimanale.xlsx")
    print("File 'orario_settimanale.xlsx' generato con successo.")
else:
    print("\nNessuna soluzione trovata.")

run_diagnostics(solver, res in (cp_model.OPTIMAL, cp_model.FEASIBLE))